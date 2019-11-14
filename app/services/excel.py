from io import BytesIO
from typing import List

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.writer.excel import save_virtual_workbook

from app.schemas.computers import ProgramInfo


def init_file_columns(worksheet: Worksheet) -> None:
    worksheet.cell(row=1, column=1).value = "NAME"
    worksheet.cell(row=1, column=2).value = "VENDOR"
    worksheet.cell(row=1, column=3).value = "VERSION"


def create_excel_file(programs_list: List[ProgramInfo], filename: str) -> BytesIO:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "new_sheet"
    init_file_columns(sheet)
    for row, program in enumerate(programs_list, 2):
        sheet.cell(row=row, column=1).value = program.name
        sheet.cell(row=row, column=2).value = program.vendor
        sheet.cell(row=row, column=3).value = program.version

    virtual_file = BytesIO(save_virtual_workbook(wb))
    virtual_file.name = filename
    return virtual_file
