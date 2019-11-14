from random import randint
from typing import Dict, List

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app import config
from app.config import PROGRAMS_EXCEL_NAME
from app.schemas.computers import ComputerItem, ProgramInfo
from app.schemas.messages import Registration


def group_computers_by_domain(computers: List[ComputerItem]) -> Dict[str, List[ComputerItem]]:  # noqa: E501
    computer_group: Dict[str, List[ComputerItem]] = {}
    for computer in computers:
        if computer.domain in computer_group:
            computer_group[computer.domain].append(computer)
        else:
            computer_group[computer.domain] = [computer]
    return computer_group


def check_group_id(event: Registration) -> bool:
    return event.group_id == config.GROUP_ID


def init_file_columns(worksheet: Worksheet) -> None:
    worksheet.cell(row=1, column=1).value = "NAME"
    worksheet.cell(row=1, column=2).value = "VENDOR"
    worksheet.cell(row=1, column=3).value = "VERSION"


def create_excel_file(programs_list: List[ProgramInfo]) -> None:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "new_sheet"
    row = 2
    col = 1
    init_file_columns(sheet)
    for program in (programs_list):
        sheet.cell(row=row, column=col).value = program.name
        sheet.cell(row=row, column=col + 1).value = program.vendor
        sheet.cell(row=row, column=col + 2).value = program.version
        row += 1
    wb.save(PROGRAMS_EXCEL_NAME)


def get_random_id() -> int:
    first_rand_value = 1
    last_rand_value = 1000000
    return randint(first_rand_value, last_rand_value)
